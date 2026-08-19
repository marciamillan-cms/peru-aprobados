"""
utils.py
--------
Formatting, filtering/search, status-label helpers, and the custom CSS
that gives the app the look of the Peru Fintech Forum website (dark,
modern, gradient-accented fintech branding) instead of a default
Streamlit dashboard.

Note on the visual design: perufintechforum.com is a Framer site without a
public style guide, so the palette below is a close, best-effort reading of
its dark, high-contrast, gradient-accented fintech aesthetic rather than a
pixel-perfect extraction of its stylesheet. Every color/spacing value is a
CSS variable at the top of THEME_CSS, so the team can fine-tune them in one
place after comparing side-by-side with the live site.
"""
from __future__ import annotations

import config
from eventtia_client import Participant

STATUS_LABELS = {
    config.STATUS_PENDING: "Nuevos",
    config.STATUS_APPROVED: "Aprobados",
    config.STATUS_REJECTED: "Rechazados",
}

STATUS_BADGE_CLASS = {
    config.STATUS_PENDING: "badge-pending",
    config.STATUS_APPROVED: "badge-approved",
    config.STATUS_REJECTED: "badge-rejected",
}


def split_by_status(participants: list[Participant]) -> dict[str, list[Participant]]:
    buckets = {config.STATUS_PENDING: [], config.STATUS_APPROVED: [], config.STATUS_REJECTED: []}
    for p in participants:
        buckets.setdefault(p.status, []).append(p)
    return buckets


def search_participants(participants: list[Participant], query: str) -> list[Participant]:
    if not query:
        return participants
    q = query.strip().lower()
    if not q:
        return participants

    def matches(p: Participant) -> bool:
        for field_name in config.SEARCHABLE_FIELDS:
            value = p.get(field_name)
            if value and q in str(value).lower():
                return True
        return False

    return [p for p in participants if matches(p)]


def field_label(field_name: str) -> str:
    return {
        "first_name": "Nombre",
        "last_name": "Apellido",
        "email": "Correo electrónico",
        "company": "Empresa",
        "job_title": "Cargo",
        "telephone": "Teléfono",
    }.get(field_name, field_name.replace("_", " ").title())


THEME_CSS = """
<style>
:root {
    /* Colors sampled directly from perufintechforum.com: deep navy
    background, vibrant emerald-green accent, white text. */
    --bg: #0E1B30;
    --surface: #16253F;
    --surface-alt: #1D2E4C;
    --border: #2C4066;
    --text: #FFFFFF;
    --text-muted: #A9B7D0;
    --accent-1: #19D882;
    --accent-2: #12B87A;
    --accent-gradient: linear-gradient(90deg, var(--accent-1), var(--accent-2));
    --success: #19D882;
    --success-bg: rgba(25, 216, 130, 0.14);
    --danger: #F5484B;
    --danger-bg: rgba(245, 72, 75, 0.14);
    --warning: #F5A623;
    --warning-bg: rgba(245, 166, 35, 0.14);
    --radius: 18px;
    --radius-sm: 999px;
    --shadow: 0 8px 24px rgba(0, 0, 0, 0.35);
}

@import url('https://fonts.googleapis.com/css2?family=Poppins:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"] {
    font-family: 'Poppins', -apple-system, BlinkMacSystemFont, sans-serif;
}

.stApp {
    background: radial-gradient(circle at 10% 0%, rgba(25, 216, 130, 0.16), transparent 40%),
                radial-gradient(circle at 90% 15%, rgba(18, 184, 122, 0.12), transparent 45%),
                var(--bg);
    color: var(--text);
}

#MainMenu, header[data-testid="stHeader"], footer {visibility: hidden;}
.block-container {padding-top: 1.5rem; max-width: 1200px;}

/* ---- Header / brand bar ---- */
.brand-bar {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 1rem 1.8rem;
    margin-bottom: 1.5rem;
    background: rgba(29, 46, 76, 0.75);
    border: 1px solid var(--border);
    border-radius: 999px;
    box-shadow: var(--shadow);
    backdrop-filter: blur(6px);
}
.brand-bar h1 {
    font-size: 1.1rem;
    font-weight: 700;
    margin: 0;
    letter-spacing: 0.01em;
}
.brand-bar .brand-mark {
    display: inline-block;
    width: 10px; height: 10px;
    border-radius: 50%;
    background: var(--accent-gradient);
    margin-right: 0.6rem;
    box-shadow: 0 0 10px rgba(25, 216, 130, 0.7);
}

/* ---- Corner logos -- fixed to the actual viewport corners, like a
   watermark, rather than sitting inside the header pill ---- */
.corner-logo {
    position: fixed;
    top: 18px;
    z-index: 999;
    height: 60px;
    max-width: 140px;
    object-fit: contain;
    filter: drop-shadow(0 2px 6px rgba(0, 0, 0, 0.4));
}
.corner-logo-left { left: 22px; }
.corner-logo-right { right: 22px; }
@media (max-width: 640px) {
    .corner-logo { height: 32px; max-width: 90px; }
    .corner-logo-left { left: 10px; }
    .corner-logo-right { right: 10px; }
}

/* ---- Participant card ---- */
.participant-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 1.1rem 1.3rem;
    margin-bottom: 0.9rem;
    box-shadow: var(--shadow);
}
.participant-card .p-name {
    font-size: 1.05rem;
    font-weight: 700;
    margin-bottom: 0.15rem;
}
.participant-card .p-role {
    color: var(--text-muted);
    font-size: 0.9rem;
    margin-bottom: 0.5rem;
}
.participant-card .p-meta {
    font-size: 0.85rem;
    color: var(--text-muted);
    line-height: 1.5;
}
.participant-card .p-meta b { color: var(--text); font-weight: 500; }

/* ---- Status badges ---- */
.badge {
    display: inline-block;
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 0.03em;
    text-transform: uppercase;
    padding: 0.22rem 0.65rem;
    border-radius: 999px;
}
.badge-pending { color: var(--warning); background: var(--warning-bg); }
.badge-approved { color: var(--success); background: var(--success-bg); }
.badge-rejected { color: var(--danger); background: var(--danger-bg); }

/* ---- Tabs ---- */
.stTabs [data-baseweb="tab-list"] { gap: 0.5rem; border-bottom: none; }
.stTabs [data-baseweb="tab-highlight"] { background-color: transparent !important; }
.stTabs [data-baseweb="tab-border"] { display: none !important; }
[role="tablist"]::after { display: none !important; }
.stTabs [data-baseweb="tab"] {
    color: var(--text-muted);
    font-weight: 600;
    padding: 0.6rem 1rem;
    border-bottom: 2px solid transparent;
}
.stTabs [aria-selected="true"] {
    color: var(--text) !important;
    border-bottom: 2px solid var(--accent-1) !important;
}

/* ---- Buttons -- fully rounded pills, matching the site's CTAs ---- */
.stButton > button {
    border-radius: 999px;
    border: 1px solid var(--border);
    background: var(--surface-alt);
    color: var(--text);
    font-weight: 600;
    padding: 0.5rem 1.3rem;
}
.stButton > button:hover {
    border-color: var(--accent-1);
    color: var(--accent-1);
}
.stButton > button[kind="primary"] {
    background: var(--accent-gradient);
    border: none;
    color: #fff;
    font-weight: 700;
}
.stButton > button[kind="primary"]:hover {
    color: #fff;
    opacity: 0.9;
}
div[data-testid="stForm"] .stButton > button {
    background: var(--accent-gradient);
    border: none;
    color: #fff;
    font-weight: 700;
}

/* ---- Download button -- same pill styling, own CSS class in Streamlit ---- */
.stDownloadButton > button {
    border-radius: 999px;
    border: 1px solid var(--accent-1);
    background: var(--surface-alt);
    color: var(--accent-1);
    font-weight: 700;
    padding: 0.5rem 1.3rem;
}
.stDownloadButton > button:hover {
    background: var(--accent-gradient);
    color: #fff;
    border-color: transparent;
}

/* ---- Inputs ---- */
.stTextInput input, .stTextInput > div > div {
    background: var(--surface-alt) !important;
    border-color: var(--border) !important;
    color: var(--text) !important;
    border-radius: 999px !important;
}

/* ---- Metric-style count chips ---- */
.count-chip {
    display: inline-block;
    font-size: 0.75rem;
    font-weight: 700;
    background: var(--surface-alt);
    border: 1px solid var(--border);
    border-radius: 999px;
    padding: 0.05rem 0.55rem;
    margin-left: 0.4rem;
}

hr { border-color: var(--border) !important; }
</style>
"""


def inject_theme(st_module):
    st_module.markdown(THEME_CSS, unsafe_allow_html=True)


def build_participants_excel(participants: list) -> bytes:
    """
    Builds a formatted .xlsx workbook (in memory, returned as bytes) with
    one row per participant, columns matching what's shown on their card.
    No formulas involved -- this is a plain data export.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Aprobados"

    columns = list(config.DISPLAY_FIELDS)
    headers = [field_label(f) for f in columns]

    header_fill = PatternFill(start_color="19D882", end_color="19D882", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="0E1B30")
    body_font = Font(name="Arial")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for row_idx, p in enumerate(participants, start=2):
        for col_idx, field_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=p.get(field_name, ""))
            cell.font = body_font

    # Reasonable column widths based on content, capped so nothing gets
    # absurdly wide from one long value.
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, len(participants) + 2):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _img_to_base64(path: str) -> str | None:
    """Reads a local image file and returns a base64 data-URI, or None if
    the file doesn't exist -- so a missing logo never crashes the app."""
    import base64
    import mimetypes
    import os

    if not path or not os.path.exists(path):
        return None
    mime, _ = mimetypes.guess_type(path)
    mime = mime or "image/png"
    with open(path, "rb") as f:
        encoded = base64.b64encode(f.read()).decode("utf-8")
    return f"data:{mime};base64,{encoded}"


def render_brand_bar(st_module, title: str, company_logo_path: str = "", event_logo_path: str = ""):
    """
    Renders the header pill (title only) plus the company/event logos
    fixed to the top-left/top-right corners of the viewport -- like a
    watermark -- rather than inside the pill itself. Either logo is
    silently skipped if its file doesn't exist.
    """
    company_logo = _img_to_base64(company_logo_path)
    event_logo = _img_to_base64(event_logo_path)

    if event_logo:
        st_module.markdown(
            f'<img src="{event_logo}" class="corner-logo corner-logo-left" alt="Logo evento" />',
            unsafe_allow_html=True,
        )
    if company_logo:
        st_module.markdown(
            f'<img src="{company_logo}" class="corner-logo corner-logo-right" alt="Logo empresa" />',
            unsafe_allow_html=True,
        )

    st_module.markdown(
        f"""
        <div class="brand-bar">
            <h1><span class="brand-mark"></span>{title}</h1>
        </div>
        """,
        unsafe_allow_html=True,
    )
